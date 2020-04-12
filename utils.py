# coding=utf-8
from openpyxl import load_workbook
from openpyxl.styles import (NamedStyle,  # 使用样式：创建命名样式
                             Font,  # 使用样式：字体，用于设置字体大小、颜色、下划线等
                             PatternFill,  # 使用样式：图样填充
    # 使用样式：边框设置
    # 使用样式：边框类型设置border_style
                             Alignment,  # 使用样式：对齐方式
    # 使用样式：保护选项
                             colors  # 颜色选项
                             )
from copy import copy

import time
import os

import config


config_path = os.getcwd() + "\\config\\excelConfig.xlsx"


def print_stdout(statement):
    print time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + " %s" % statement


def cell_to_str(cell):
    return unicode.strip(cell.value)


def col_to_set(col):
    res = set()
    for i in range(len(col)):
        if i == 0:
            continue
        if col[i].value == '#':
            break
        res.add(unicode.strip(col[i].value))
    return res


def col_to_list(col):
    res = []
    for i in range(len(col)):
        if i == 0:
            continue
        if col[i].value == '#':
            break
        res.append(unicode.strip(col[i].value))
    return res


def col_to_output_config(col):
    res = set()
    res_key = []
    for i in range(len(col)):
        if i == 0:
            continue
        val = col[i].value
        if val == '#':
            break
        if val.startswith('%'):
            res_key.append(unicode.lstrip(val, '%'))
            continue
        res.add(unicode.strip(col[i].value))
    return res, res_key


def cell_include_check(target, incl):
    if not incl or len(incl) == 0:
        return True
    if not target or target == "":
        return False
    for word in incl:
        if word in target:
            return True
    return False


def cell_exclude_check(target, excl):
    if not excl or len(excl) == 0:
        return True
    if not target or target == "":
        return True
    for w in excl:
        if w in target:
            return False
    return True


def _get_data_filter_excel(cfg):
    cfg_dict = {'read_path': cell_to_str(cfg['A2']), 'output_path': cell_to_str(cfg['B2']),
                'match': col_to_set(cfg['C']),
                'include': col_to_set(cfg['D']), 'exclude': col_to_set(cfg['E']),
                'check_column': cell_to_str(cfg['F2']),
                'output_columns': col_to_output_config(cfg['G'])[0],
                'output_keyword': col_to_output_config(cfg['G'])[1],
                'titles': col_to_list(cfg['H'])}
    title_style = cfg['I2']
    cfg_dict['title_style'] = title_style
    cfg_dict['style_titleRow'] = NamedStyle(name='style_titleRow',
                                            font=copy(title_style.font),
                                            fill=copy(title_style.fill),
                                            alignment=copy(title_style.alignment)
                                            )
    return cfg_dict


def _get_keyword_count_excel(cfg):
    cfg_dict = {'read_path': cell_to_str(cfg['A2']), 'output_path': cell_to_str(cfg['B2']),
                'keywords': col_to_set(cfg['C']), 'output_columns': col_to_set(cfg['D'])}
    return cfg_dict


def load_config(sheet_name=None):
    if not os.path.exists(config_path):
        print_stdout('Error: can not find config file, please check! [%s]' % config_path)
        return
    cfg_dict = {}
    if not sheet_name:
        sheet = load_workbook(filename=config_path).active
        cfg_dict = _get_data_filter_excel(sheet)
    else:
        sheet = load_workbook(filename=config_path).get_sheet_by_name(sheet_name)
        cfg_dict = _get_keyword_count_excel(sheet)

    return config.Config(cfg_dict)

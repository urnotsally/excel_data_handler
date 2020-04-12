from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.utils import column_index_from_string

import os
import sys
import utils
cfg = utils.load_config('keywordcount')


def keyword_count(row):
    key_match = {}
    out_row = []
    match_row = []
    row_index = None
    for cell in row:
        if isinstance(cell, EmptyCell):
            continue
        value = cell.value
        if not isinstance(value, unicode) and not isinstance(value, str):
            continue
        if cell.column_letter in cfg.output_columns:
            out_row.append(value)
        for key in cfg.keywords:
            if key in value:
                last = key_match.get(key, None)
                if not last:
                    key_match[key] = value
                else:
                    if last != value:
                        match_row.append(last)
                        match_row.append(value)
                        row_index = cell.row
                break
    if len(match_row) > 1:
        utils.print_stdout('output row' + row_index + "...")
        return out_row + match_row
    return []


def main():
    utils.print_stdout('Load Config...')
    read_path = cfg.read_path
    if not os.path.exists(read_path):
        utils.print_stdout('Error: read_path[%s] is not exists, please check!' % read_path)
        return

    wb1 = Workbook()
    out_sheet = wb1.active

    utils.print_stdout('->start data loading... [%s]' % read_path)
    from_sheet = load_workbook(filename=read_path, read_only=True).active
    utils.print_stdout('Data Loading Done')
    rows = from_sheet.rows

    for row in rows:
        res = keyword_count(row)
        if len(res) > 1:
            out_sheet.append(res)

    wb1.save(cfg.output_path)


if __name__ == '__main__':
    sys.exit(main())
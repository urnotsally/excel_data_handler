# coding=utf-8
import os
import sys

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.utils import column_index_from_string

import utils


cfg = utils.load_config('datafilter')
setTitleStyle = True

def format_row(row):
    new_raw = []
    for cell in row:
        if isinstance(cell, EmptyCell):
            continue
        if cell.column_letter in cfg.output_columns:
            new_raw.append(cell.value)
            continue
        val = cell.value
        if not isinstance(val, unicode) and not isinstance(val, str):
            continue
        for keyword in cfg.output_keyword:
            if keyword in val:
                new_raw.append(val)
                break
    return new_raw


def main():
    read_path = cfg.read_path
    check_column = cfg.check_column
    utils.print_stdout('Load Config...')

    if not os.path.exists(read_path):
        utils.print_stdout('Error: read_path[%s] is not exists, please check!' % read_path)
        return

    wb1 = Workbook()
    passenger_sheet = wb1.active
    passenger_sheet.title = "passenger_car"
    other_sheet = wb1.create_sheet(title="others")
    if cfg.titles and len(cfg.titles) != 0:
        utils.print_stdout("->setting titles...")
        if len(cfg.titles) != len(cfg.output_columns)+len(cfg.output_keyword):
            cfg.print_stdout('Error: the length of titles is different from output_column\'s, please check!')
            return 1
        passenger_sheet.append(cfg.titles)
        other_sheet.append(cfg.titles)
        if setTitleStyle:
            for cell in passenger_sheet[1]:
                cell.style = cfg.style_titleRow
            for cell in other_sheet[1]:
                cell.style = cfg.style_titleRow

    utils.print_stdout('->start data loading... [%s]' % read_path)
    wb2 = load_workbook(filename=read_path, read_only=True)
    utils.print_stdout('Data Loading Done')

    data = wb2.active
    rows = data.rows
    check_col = column_index_from_string(check_column)

    utils.print_stdout('->start data filtering...')

    for row in rows:
        value = row[check_col - 1].value
        if not isinstance(value, unicode) and not isinstance(value, str):
            utils.print_stdout('value type is not unicode or str, so ignore value [%s]' % value)
            continue
        if value in cfg.match:
            utils.print_stdout('value match [%s]' % value)
            new_raw = format_row(row)
            passenger_sheet.append(new_raw)
            continue
        if utils.cell_include_check(value, cfg.include) and utils.cell_exclude_check(value, cfg.exclude):
            utils.print_stdout('value include and exclude [%s]' % value)
            new_raw = format_row(row)
            other_sheet.append(new_raw)

    utils.print_stdout('Data Filtration Done')

    wb1.save(filename=cfg.output_path)
    utils.print_stdout('Save result to Excel [%s]' % cfg.output_path)


if __name__ == '__main__':
    sys.exit(main())

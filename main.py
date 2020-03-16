# coding=utf-8
import os
import sys
import time
from copy import copy

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.styles import (NamedStyle,  # 使用样式：创建命名样式
                             Font,  # 使用样式：字体，用于设置字体大小、颜色、下划线等
                             PatternFill,  # 使用样式：图样填充
    # 使用样式：边框设置
    # 使用样式：边框类型设置border_style
                             Alignment,  # 使用样式：对齐方式
    # 使用样式：保护选项
                             colors  # 颜色选项
                             )
from openpyxl.utils import column_index_from_string

config_path = os.getcwd() + "\\ExcelFilterConfig.xlsx"

# setting
read_path = r'D:\202003data.xlsx'
output_path = r'D:\output.xlsx'

# used in passenger car
match = {u"插电式混合动力多用途乘用车", u"纯电动多用途乘用车", u"插电式混合动力轿车", u"纯电动轿车", u"插电式增程混合动力轿车", u"纯电动运动型乘用车"}
# used in other car
include = {u'电'}
exclude = {u"摩托", u"客车"}
# specify column in origin excl
check_column = 'M'
output_columns = {'C', 'E', 'G', 'I', 'AI'}
output_keyword = ['储能装置种类']
titles = [u'标题1', u'标题2', u'标题3', u'标题4', u'标题5']
setTitleStyle = True
style_titleRow = NamedStyle(name='style_titleRow',
                            font=Font(b=True),  # 粗体
                            fill=PatternFill(fill_type='solid',  # 指定填充的类型，支持的有：'solid'等。
                                             fgColor=colors.DARKGREEN  # 指定填充颜色
                                             ),
                            alignment=Alignment(horizontal='center',  # 水平居中
                                                vertical='center',  # 垂直居中
                                                wrap_text=True,  # 自动换行
                                                )
                            )


def load_config():
    if not os.path.exists(config_path):
        print_stdout('Error: can not find config file, please check! [%s]' % config_path)
        return
    cfg = load_workbook(filename=config_path).active
    global read_path, output_path, match, include, exclude, check_column, output_columns, titles, style_titleRow, output_keyword
    read_path = cell_to_str(cfg['A2'])
    output_path = cell_to_str(cfg['B2'])
    match = col_to_set(cfg['C'])
    include = col_to_set(cfg['D'])
    exclude = col_to_set(cfg['E'])
    check_column = cell_to_str(cfg['F2'])
    output_columns, output_keyword = col_to_output_config(cfg['G'])
    titles = col_to_list(cfg['H'])
    title_style = cfg['I2']
    style_titleRow = NamedStyle(name='style_titleRow',
                                font=copy(title_style.font),
                                fill=copy(title_style.fill),
                                alignment=copy(title_style.alignment)
                                )
    return


def cell_to_str(cell):
    return unicode.strip(cell.value)


def col_to_set(col):
    res = set()
    for i in range(len(col)):
        if i == 0:
            continue
        if col[i].value == '#':
            break
        res.add(unicode.strip(col[i].value))
    return res


def col_to_list(col):
    res = []
    for i in range(len(col)):
        if i == 0:
            continue
        if col[i].value == '#':
            break
        res.append(unicode.strip(col[i].value))
    return res


def col_to_output_config(col):
    res = set()
    res_key = []
    for i in range(len(col)):
        if i == 0:
            continue
        val = col[i].value
        if val == '#':
            break
        if val.startswith('%'):
            res_key.append(unicode.lstrip(val, '%'))
            continue
        res.add(unicode.strip(col[i].value))
    return res, res_key


def include_check(target, incl):
    if not incl or len(incl) == 0:
        return True
    if not target or target == "":
        return False
    for word in incl:
        if word in target:
            return True
    return False


def exclude_check(target, excl):
    if not excl or len(excl) == 0:
        return True
    if not target or target == "":
        return True
    for w in excl:
        if w in target:
            return False
    return True


def print_stdout(statement):
    print time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + " %s" % statement


def format_row(row):
    new_raw = []
    for cell in row:
        if isinstance(cell, EmptyCell):
            continue
        if cell.column_letter in output_columns:
            new_raw.append(cell.value)
            continue
        val = cell.value
        if not isinstance(val, unicode) and not isinstance(val, str):
            continue
        for keywory in output_keyword:
            if keywory in val:
                new_raw.append(val)
                break
    return new_raw


def main():
    print_stdout('Load Config...')
    load_config()

    if not os.path.exists(read_path):
        print_stdout('Error: read_path[%s] is not exists, please check!' % read_path)
        return

    wb1 = Workbook()
    passenger_sheet = wb1.active
    passenger_sheet.title = "passenger_car"
    other_sheet = wb1.create_sheet(title="others")
    if titles and len(titles) != 0:
        print_stdout("->setting titles...")
        if len(titles) != len(output_columns)+len(output_keyword):
            print_stdout('Error: the length of titles is different from output_column\'s, please check!')
            return 1
        passenger_sheet.append(titles)
        other_sheet.append(titles)
        if setTitleStyle:
            for cell in passenger_sheet[1]:
                cell.style = style_titleRow
            for cell in other_sheet[1]:
                cell.style = style_titleRow

    print_stdout('->start data loading... [%s]' % read_path)
    wb2 = load_workbook(filename=read_path, read_only=True)
    print_stdout('Data Loading Done')

    data = wb2.active
    rows = data.rows
    check_col = column_index_from_string(check_column)

    print_stdout('->start data filtering...')

    for row in rows:
        value = row[check_col - 1].value
        if not isinstance(value, unicode) and not isinstance(value, str):
            print_stdout('value type is not unicode or str, so ignore value [%s]' % value)
            continue
        if value in match:
            print_stdout('value match [%s]' % value)
            new_raw = format_row(row)
            passenger_sheet.append(new_raw)
            continue
        if include_check(value, include) and exclude_check(value, exclude):
            print_stdout('value include and exclude [%s]' % value)
            new_raw = format_row(row)
            other_sheet.append(new_raw)

    print_stdout('Data Filtration Done')

    wb1.save(filename=output_path)
    print_stdout('Save result to Excel [%s]' % output_path)


if __name__ == '__main__':
    sys.exit(main())
